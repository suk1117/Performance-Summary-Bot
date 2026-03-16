import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "포트폴리오"

# 헤더 (dashboard.py가 읽는 필수 컬럼)
headers    = ["종목명", "티커", "국가", "비중(%)", "평단가", "통화"]
col_widths = [16, 12, 8, 10, 16, 8]

header_fill = PatternFill(start_color="0D1526", end_color="0D1526", fill_type="solid")
header_font = Font(color="00D4FF", bold=True, size=11)
thin = Border(
    left=Side(style='thin', color="1E2D4A"),
    right=Side(style='thin', color="1E2D4A"),
    top=Side(style='thin', color="1E2D4A"),
    bottom=Side(style='thin', color="1E2D4A"),
)

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 26

# ── 예시 데이터 ──────────────────────────────────────────────────
# 컬럼: 종목명 / 티커 / 국가 / 비중(%) / 평단가 / 통화
#
# 국가:  KR   → pykrx 조회  (티커: 6자리 숫자, 예: 005930)
#        US   → yfinance 조회 (티커: 영문, 예: AAPL)
#        현금  → 수익률 0 고정
#
# 통화:  KRW (원화 기준 평단가)
#        USD (달러 기준 평단가 → 대시보드에서 그대로 표시)
#
# 비중(%) 합계는 100이 되도록 맞추세요.
examples = [
    # 종목명         티커       국가   비중   평단가      통화
    ["삼성전자",    "005930",  "KR",  25.0,  72000,    "KRW"],
    ["SK하이닉스",  "000660",  "KR",  15.0,  130000,   "KRW"],
    ["애플",        "AAPL",    "US",  20.0,  175.00,   "USD"],
    ["엔비디아",    "NVDA",    "US",  15.0,  480.00,   "USD"],
    ["KODEX 200",  "069500",  "KR",  10.0,  35000,    "KRW"],
    ["현금",        "CASH",    "현금", 15.0, 1000000,  "KRW"],
]

even_fill = PatternFill(start_color="0F1A2E", end_color="0F1A2E", fill_type="solid")
odd_fill  = PatternFill(start_color="0D1526", end_color="0D1526", fill_type="solid")
text_font = Font(color="E2E8F0", size=10)

for row_idx, row_data in enumerate(examples, 2):
    fill = even_fill if row_idx % 2 == 0 else odd_fill
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.font = text_font
        cell.border = thin
        cell.alignment = Alignment(
            horizontal="right" if col_idx in [4, 5] else "center"
        )

# ── 안내 시트 ────────────────────────────────────────────────────
ws2 = wb.create_sheet("작성 안내")
guide = [
    ["📋 포트폴리오 작성 안내"],
    [],
    ["컬럼",       "설명"],
    ["종목명",     "자유롭게 입력 (예: 삼성전자, 애플, 현금)"],
    ["티커",       "KR: 6자리 종목코드 (005930) / US: 영문 티커 (AAPL) / 현금: CASH"],
    ["국가",       "KR / US / 현금 중 하나 입력 (대소문자 구분)"],
    ["비중(%)",    "포트폴리오 내 비중. 합계가 100이 되도록 입력"],
    ["평단가",     "KRW 통화면 원화 금액, USD면 달러 금액 입력"],
    ["통화",       "KRW 또는 USD"],
    [],
    ["⚠️ 주의사항", ""],
    ["",           "1행(헤더)은 수정하지 마세요"],
    ["",           "비중(%) 합계를 100으로 맞추세요"],
    ["",           "국가 값은 정확히 KR / US / 현금 으로 입력"],
    ["",           "현금 항목은 수익률이 0%로 표시됩니다"],
    [],
    ["💡 KR 티커 찾기",  "https://finance.naver.com 검색 → URL의 숫자 6자리"],
    ["💡 US 티커 찾기",  "https://finance.yahoo.com 검색 → 심볼(Symbol)"],
]

for r, row in enumerate(guide, 1):
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = Font(size=11 if r == 1 else 10, bold=(r in [1, 3]))

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 60

wb.save("portfolio.xlsx")
print("portfolio.xlsx 생성 완료!")
print("   -> 파일 열어서 예시 데이터 지우고 본인 포트폴리오 입력하세요.")
print()
print("   필수 컬럼: 종목명 / 티커 / 국가 / 비중(%) / 평단가 / 통화")
print("   국가 값:   KR (한국주식) / US (미국주식) / 현금")
